"""
Exports the current session's rooms to a multi-sheet Excel workbook - one
sheet per working tab (Room Schedule, HVAC & FCU, Ventilation, Water
Services, Heat Load), so "export the whole platform" gives one file with
everything, and "export an individual tab" is just opening/printing the
one sheet you want from it. The auditable, client-issuable deliverable,
generated from whatever is currently in the Streamlit tables.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import calc_engine
import reference_data as ref

NAVY = "1B365D"
ICE_BLUE = "E6F0FA"
HEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TOTAL_FILL = PatternFill("solid", fgColor=ICE_BLUE)
DISCLAIMER_TEXT = (
    "Disclaimer: this platform is a calculation aid only and does not hold or assume any design "
    "liability. All figures must be independently checked and verified by a suitably qualified "
    "engineer against current standards and project-specific requirements before use for design, "
    "construction, or compliance purposes."
)


def _write_sheet(wb, title, headers, rows, totals_row=None, col_widths=None):
    """Writes one styled sheet: header row, data rows, an optional TOTALS
    row, and the disclaimer footer - shared by every sheet in this export
    so they all look consistent."""
    ws = wb.create_sheet(title=title)
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    r = 2
    for row in rows:
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=value)
        r += 1

    if totals_row is not None:
        for col, value in enumerate(totals_row, start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.fill = TOTAL_FILL
        r += 1

    n_cols = len(headers)
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

    disclaimer_row = r + 1
    ws.cell(row=disclaimer_row, column=1, value=DISCLAIMER_TEXT)
    ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=n_cols)
    ws.cell(row=disclaimer_row, column=1).font = Font(name="Segoe UI", size=9, italic=True, color="C0392B")
    ws.cell(row=disclaimer_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[disclaimer_row].height = 30
    return ws


def build_export_workbook(rooms: list, fresh_air_rate_ls_person: float = None) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)  # remove the default blank sheet - every sheet here is created explicitly

    catalogue = ref.FCU_CATALOGUE
    lu_values = ref.FIXTURE_LU

    # ---- Room Schedule ----
    schedule_rows = [
        (r.get("name"), r.get("floor"), r.get("area_m2"), r.get("ceiling_height_m"),
         r.get("summer_design_temp_c"), r.get("winter_design_temp_c"),
         "Yes" if r.get("include_in_summary", True) else "No")
        for r in rooms
    ]
    _write_sheet(
        wb, "Room Schedule",
        ["Room Name", "Floor", "Area (m2)", "Ceiling Height (m)", "Summer Temp (C)", "Winter Temp (C)",
         "Include in Print Summary"],
        schedule_rows,
        col_widths={"A": 26, "B": 10, "C": 12, "D": 14, "E": 14, "F": 14, "G": 18},
    )

    # ---- HVAC & FCU ----
    hvac_rows = []
    total_sensible = total_latent = total_load = 0.0
    for room in rooms:
        gains = calc_engine.calculate_heat_gains(room)
        if gains.is_uncontrolled:
            fcu = None
            fcu_model, fcu_qty, fcu_status = "Not Required (Uncontrolled)", "-", "Uncontrolled"
        else:
            fcu = calc_engine.select_fcu(
                gains.total_cooling_load_kw, room.get("manufacturer", "Daikin"),
                room.get("unit_type", "Ducted"), room.get("quantity", 1), catalogue,
            )
            fcu_model = fcu.selected_model if fcu else "-"
            fcu_qty = room.get("quantity", 1) if fcu else "-"
            fcu_status = "TBC" if (fcu and fcu.is_tbc) else (("PASS" if fcu.meets_load else "REVIEW") if fcu else "-")
        hvac_rows.append((
            room.get("name"), gains.volume_m3, gains.design_temp_c, gains.total_sensible_kw,
            gains.total_latent_kw, gains.total_cooling_load_kw,
            fcu_model, fcu_qty, fcu_status,
        ))
        total_sensible += gains.total_sensible_kw
        total_latent += gains.total_latent_kw
        total_load += gains.total_cooling_load_kw
    _write_sheet(
        wb, "HVAC & FCU",
        ["Room Name", "Volume (m3)", "Summer Design Temp (C)", "Sensible (kW)", "Latent (kW)",
         "Total Load (kW)", "Selected FCU", "Qty", "Status"],
        hvac_rows,
        totals_row=("TOTALS", None, None, round(total_sensible, 2), round(total_latent, 2), round(total_load, 2), None, None, None),
        col_widths={"A": 26, "B": 12, "C": 16, "D": 12, "E": 12, "F": 14, "G": 16, "H": 8, "I": 10},
    )

    # ---- Ventilation ----
    vent_rows = []
    total_airflow = 0.0
    for room in rooms:
        gains = calc_engine.calculate_heat_gains(room)
        vent = calc_engine.calculate_ventilation(room, gains.volume_m3, fresh_air_rate_ls_person)
        vent_rows.append((room.get("name"), vent.required_design_airflow_ls, vent.selected_duct_size_mm))
        total_airflow += vent.required_design_airflow_ls
    _write_sheet(
        wb, "Ventilation",
        ["Room Name", "Required Airflow (l/s)", "Selected Duct Size (mm)"],
        vent_rows,
        totals_row=("TOTALS", round(total_airflow, 1), None),
        col_widths={"A": 26, "B": 20, "C": 20},
    )

    # ---- Water Services ----
    water_rows = []
    total_lu = 0.0
    for room in rooms:
        water = calc_engine.calculate_room_loading_units(room, lu_values)
        water_rows.append((room.get("name"), water.loading_units))
        total_lu += water.loading_units
    _write_sheet(
        wb, "Water Services",
        ["Room Name", "Loading Units (LU)"],
        water_rows,
        totals_row=("TOTALS", round(total_lu, 1)),
        col_widths={"A": 26, "B": 16},
    )

    # ---- Heat Load (Winter) ----
    heatload_rows = []
    total_heatloss = 0.0
    for room in rooms:
        gains = calc_engine.calculate_heat_gains(room)
        heatloss = calc_engine.calculate_winter_heat_loss(room, gains.volume_m3)
        heatload_rows.append((
            room.get("name"), heatloss.fabric_loss_w, heatloss.infiltration_loss_w, heatloss.total_heat_loss_kw,
        ))
        total_heatloss += heatloss.total_heat_loss_kw
    _write_sheet(
        wb, "Heat Load (Winter)",
        ["Room Name", "Fabric Loss (W)", "Infiltration Loss (W)", "Total Heat Loss (kW)"],
        heatload_rows,
        totals_row=("TOTALS", None, None, round(total_heatloss, 2)),
        col_widths={"A": 26, "B": 14, "C": 18, "D": 18},
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_revit_csv(rooms: list, fresh_air_rate_ls_person: float = None) -> str:
    """A clean, simple CSV designed specifically for a Revit/pyRevit
    import script to read - deliberately plain (no styling, no merged
    cells, no multi-sheet structure) since that's much easier and more
    robust for a script to parse than the full styled .xlsx export.
    Matched to Revit Rooms by Room Name - see the accompanying pyRevit
    script (revit_import_mep_data.py) for the Revit-side half of this."""
    import csv
    import io as io_module

    rows = []
    for room in rooms:
        gains = calc_engine.calculate_heat_gains(room)
        vent = calc_engine.calculate_ventilation(room, gains.volume_m3, fresh_air_rate_ls_person)
        if gains.is_uncontrolled:
            fcu = None
            fcu_model, fcu_status = "Not Required (Uncontrolled)", "Uncontrolled"
        else:
            fcu = calc_engine.select_fcu(
                gains.total_cooling_load_kw, room.get("manufacturer", "Daikin"),
                room.get("unit_type", "Ducted"), room.get("quantity", 1), ref.FCU_CATALOGUE,
            )
            fcu_model = fcu.selected_model if fcu else "No Suitable Unit"
            fcu_status = "TBC" if (fcu and fcu.is_tbc) else (("PASS" if fcu.meets_load else "REVIEW") if fcu else "-")
        heatloss = calc_engine.calculate_winter_heat_loss(room, gains.volume_m3)
        water = calc_engine.calculate_room_loading_units(room)
        grille = calc_engine.select_grille_diffuser(
            vent.required_design_airflow_ls, room.get("grille_type") or ref.GRILLE_TYPES[0],
            ref.GRILLE_DIFFUSER_CATALOGUE, room.get("grille_qty", 1),
        )

        rows.append({
            "Room Name": room.get("name", ""),
            "Area (m2)": room.get("area_m2", 0),
            "Volume (m3)": gains.volume_m3,
            "Sensible Load (kW)": gains.total_sensible_kw,
            "Latent Load (kW)": gains.total_latent_kw,
            "Total Cooling Load (kW)": gains.total_cooling_load_kw,
            "Selected FCU": fcu_model,
            "FCU Status": fcu_status,
            "Required Airflow (l/s)": vent.required_design_airflow_ls,
            "Duct Size (mm)": vent.selected_duct_size_mm,
            "Grille/Diffuser Type": grille.grille_type if grille else "-",
            "Grille/Diffuser Qty": grille.quantity if grille else "-",
            "Grille/Diffuser Size": grille.size if grille else "-",
            "Winter Heat Loss (kW)": heatloss.total_heat_loss_kw,
            "Loading Units (LU)": water.loading_units,
        })

    output = io_module.StringIO()
    if rows:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return output.getvalue()
